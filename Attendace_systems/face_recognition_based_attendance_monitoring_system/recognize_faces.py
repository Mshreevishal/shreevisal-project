import cv2
import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


def recognize():

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read("trainer/model.yml")

    faceCascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    )

    cam = cv2.VideoCapture(0)

    font = cv2.FONT_HERSHEY_SIMPLEX

    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()

    recorded_ids = set()

    while True:

        ret, img = cam.read()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale(gray, 1.2, 5)

        for (x, y, w, h) in faces:

            student_id, conf = recognizer.predict(gray[y:y+h, x:x+w])

            if conf < 60:

                cursor.execute(
                    "SELECT name FROM students WHERE id=?",
                    (student_id,)
                )

                result = cursor.fetchone()

                if result:
                    name = result[0]
                else:
                    name = "Unknown"

                now = datetime.now()
                date = now.strftime("%Y-%m-%d")
                time = now.strftime("%H:%M:%S")

                if student_id not in recorded_ids:

                    file_name = "attendance.xlsx"

                    if not os.path.exists(file_name):

                        wb = Workbook()
                        sheet = wb.active
                        sheet.append(["ID","Name","Date","Time"])

                    else:

                        wb = load_workbook(file_name)
                        sheet = wb.active

                    sheet.append([student_id,name,date,time])
                    wb.save(file_name)

                    recorded_ids.add(student_id)

                label = name

            else:
                label = "Unknown"

            cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),2)

            cv2.putText(img,label,(x,y-10),
                        font,0.9,(255,255,255),2)

        cv2.imshow("Face Recognition Attendance",img)

        if cv2.waitKey(1)==ord('q'):
            break

    cam.release()
    cv2.destroyAllWindows()
    conn.close()